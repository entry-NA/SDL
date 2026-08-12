# -*- coding: utf-8 -*-
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\23991\Desktop\实验数据.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = [str(v) if v is not None else '' for v in row]
        print('\t'.join(vals))